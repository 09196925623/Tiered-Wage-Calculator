#region ///////////////////////////////////////////////////////////// libraries
from tkinter import *
from tkinter import messagebox, ttk, filedialog
import webbrowser
import os
import subprocess
import shutil
import platform
try:
    import pyperclip
    HAS_PYPERCLIP = True
except ImportError:
    HAS_PYPERCLIP = False
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except Exception:
    # openpyxl is optional at runtime; set names to None so
    # linter/runtime checks won't complain about undefined names.
    HAS_OPENPYXL = False
    Workbook = None
    load_workbook = None
    Font = None
    PatternFill = None
    Alignment = None
    Border = None
    Side = None

import sys
import os

def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

#endregion
#region ///////////////////////////////////////////////////////////// discription
#Creat and design by Amin Ahankoubi
#E-mail : amin.ahankobi@gmail.com
#endregion
#region ///////////////////////////////////////////////////////////// root
root = Tk()
root.title("Tiered Wage V.1.2.0")
root.geometry("500x800")
root.minsize(400, 500)  # حداقل اندازه پنجره را کاهش دادم
root.resizable(False, False)  # فعال کردن resize
root.iconbitmap(resource_path("coin.ico"))

# رنگ‌های مدرن و بروز
COLORS = {
    'bg': '#F8FAFC',
    'card': '#FFFFFF',
    'accent': '#6366F1',
    'accent_hover': '#4F46E5',
    'accent_light': '#EEF2FF',
    'text': '#0F172A',
    'text_secondary': '#475569',
    'text_light': '#94A3B8',
    'border': '#E2E8F0',
    'border_light': '#F1F5F9',
    'success': '#10B981',
    'success_light': '#D1FAE5',
    'error': '#EF4444',
    'error_light': '#FEE2E2',
    'header': '#1E293B',
}

root.configure(bg=COLORS['bg'])

# تنظیم grid برای layout بهتر
root.grid_rowconfigure(0, weight=0)  # Header
root.grid_rowconfigure(1, weight=1)  # Scrollable content
root.grid_columnconfigure(0, weight=1)
#endregion
#region ///////////////////////////////////////////////////////////// variables
font_title = "Segoe UI", "18", "bold"
font_heading = "Segoe UI", "10", "bold"
font_body = "Segoe UI", "10"
font_small = "Segoe UI", "9"
font_button = "Segoe UI", "11", "bold"

padding = 10
inner_padding = 7
card_radius = 8

url102 = "https://pii.ir/wp-content/uploads/2022/04/%D8%A2%DB%8C%DB%8C%D9%86-%D9%86%D8%A7%D9%85%D9%87-102.pdf"
url102_1 = "https://www.centinsur.ir/_douranportal/documents/2277/102-1.pdf"

bzyLevel1 = int(0)
bzyLevel2 = int(18000000000)
bzyLevel3 = int(36000000000)
bzyLevel4 = int(72000000000)

exptlevel1 = int(0)
exptlevel2 = int(3600000000)
exptlevel3 = int(18000000000)
exptlevel4 = int(36000000000)

# متغیر برای ذخیره مسیر فایل اکسل ورودی
excel_input_path = None
#endregion
#region ///////////////////////////////////////////////////////////// functions
def centerWindow(): #بازکردن پنجره در وسط صفحه
    root.update_idletasks()

    w = root.winfo_width()
    h = root.winfo_height()

    ws = root.winfo_screenwidth()
    hs = root.winfo_screenheight()

    x = (ws/2) - (w/2)
    y = (hs/2) - (h/2)
    root.geometry('%dx%d+%d+%d' %(w,h,x,y))

def openWeb1(): #دانلود لینک 1
    webbrowser.open(url102)

def openWeb2(): #دانلود لینک 2
    webbrowser.open(url102_1)
    
def aboutUs(): #نمایش درباره ما
    messagebox.showinfo("         درباره ما         ",
    "این نرم افزار صرفا جهت محاسبه کارمزد پلکانی طبق آئین نامه بیمه مرکزی طراحی شده و تمامی حقوق نشر و نصب آن متعلق به شرکت کاریزمابیمه می باشد\n\n"
    "Software Developer : Amin Ahankoubi\n\n"
    "E-mail : amin.ahankobi@gmail.com\n"
    "NWdev from PNLdev.com team"
    ) 
    
def showInfo(): #نمایش راهنما
    messagebox.showinfo("               راهنما               " ,
    "1-با توجه به قوانین آئین نامه و تاریخ ابلاغ و اجرا بخشنامه ها، متناسب با تاریخ صدور بیمه نامه خود، مدل محاسبه پلکانی را انتخاب کنید.\n\n"
    "2-درصد بازاریابی و صدور را طبق آئین نامه ثبت کرده و مبلغ حق بیمه قابل پرداخت را وارد کنید تا مبلغ 'کارمزد از حق بیمه' محاسبه شود.\n\n"
    "3-در صورت محسابه کارمزد بازاریاب طبق توافق با ایشان، نوع 'سهم بازاریاب از کارمزد' را انتخاب کنید.\n\n"
    "4-در صورت خصوصی یا دولتی بودن بیمه گزار، از بخش 'خصوصی/دولتی' آن را انتخاب کنید.\n\n"
    "5-در صورت نیاز به محاسبه کارمزد از مبلغ حق بیمه خالص، یکی از گزینه های 'کسر مالیات' را انتخاب کنید.\n\n\n"
    "* در بخش کسر مبلغ کارمزد تنها 'یک گزینه' را انتخاب کنید"
    )

def clear_error_message(): #برای پاک کردن متن خطای قبلی
    labelErrors.config(text="")

def copy_to_clipboard(text): #کپی کردن متن به clipboard
    """کپی کردن متن به clipboard"""
    try:
        if HAS_PYPERCLIP:
            pyperclip.copy(text)
        else:
            # استفاده از clipboard داخلی tkinter
            root.clipboard_clear()
            root.clipboard_append(text)
            root.update()
        # نمایش پیام موفقیت
        labelErrors.config(text="✓ نتیجه به clipboard کپی شد!", fg=COLORS['success'])
        root.after(3000, lambda: labelErrors.config(text="") if labelErrors.cget("text") == "✓ نتیجه به clipboard کپی شد!" else None)
    except Exception as e:
        messagebox.showerror("خطا", f"خطا در کپی کردن: {str(e)}")

def copy_result(): #کپی کردن نتیجه
    """کپی کردن نتیجه محاسبه"""
    result_text = labelFinalKarmozd.cget("text")
    if result_text and result_text != "0 ریال":
        # حذف "ریال" و کپی کردن فقط عدد
        number = result_text.replace(" ریال", "").replace(",", "")
        copy_to_clipboard(number)
    else:
        messagebox.showwarning("هشدار", "ابتدا محاسبه را انجام دهید!")

def open_excel_file(file_path): #باز کردن فایل اکسل
    """باز کردن فایل اکسل با برنامه پیش‌فرض"""
    try:
        if platform.system() == 'Windows':
            os.startfile(file_path)
        elif platform.system() == 'Darwin':  # macOS
            subprocess.call(['open', file_path])
        else:  # Linux
            subprocess.call(['xdg-open', file_path])
    except Exception as e:
        messagebox.showerror("خطا", f"خطا در باز کردن فایل: {str(e)}")

def download_template_file():
    """Copy the bundled `template.xlsx` next to the project to a user-selected path."""
    try:
        src = resource_path('template.xlsx')
        if not os.path.exists(src):
            messagebox.showerror('خطا', 'فایل نمونه (template.xlsx) در پوشهٔ پروژه یافت نشد.')
            return
        dest = filedialog.asksaveasfilename(
            defaultextension='.xlsx',
            initialfile='template.xlsx',
            filetypes=[('Excel files', '*.xlsx'), ('All files', '*.*')],
            title='ذخیره نمونه فایل'
        )
        if not dest:
            return
        shutil.copyfile(src, dest)
        messagebox.showinfo('موفق', f'نمونه فایل ذخیره شد.\n\nمسیر: {dest}')
        # Optionally open it
        open_excel_file(dest)
    except Exception as e:
        messagebox.showerror('خطا', f'خطا در ذخیره نمونه فایل:\n{str(e)}')

def create_excel_template(): #ایجاد template اکسل ورودی
    """ایجاد فایل template اکسل برای ورودی"""
    global HAS_OPENPYXL, Workbook, load_workbook, Font, PatternFill, Alignment, Border, Side
    if not HAS_OPENPYXL:
        result = messagebox.askyesno(
            "کتابخانه مورد نیاز", 
            "کتابخانه openpyxl برای کار با فایل‌های اکسل نیاز است.\n\n"
            "آیا می‌خواهید اکنون نصب شود؟\n\n"
            "(نیاز به دسترسی به اینترنت)"
        )
        if result:
            try:
                import subprocess
                import sys
                subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
                messagebox.showinfo("موفق", "کتابخانه openpyxl با موفقیت نصب شد!\nلطفا دوباره تلاش کنید.")
                try:
                    global Workbook, load_workbook, Font, PatternFill, Alignment, Border, Side
                    from openpyxl import Workbook, load_workbook
                    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                    HAS_OPENPYXL = True
                except Exception:
                    pass
            except Exception as e:
                messagebox.showerror("خطا", f"خطا در نصب کتابخانه:\n{str(e)}\n\nلطفا دستی نصب کنید:\npip install openpyxl")
        return None
    
    try:
        # انتخاب مسیر ذخیره
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            title="ذخیره فایل ورودی اکسل"
        )
        
        if not file_path:
            return None
        
        wb = Workbook()
        ws = wb.active
        ws.title = "ورودی"
        
        # استایل برای هدر
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")
        border_style = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # نوشتن هدرها
        headers = [
            "ردیف",
            "کارمزد بازاریابی (%)",
            "درصد صدور (%)",
            "سهم بازاریاب از کارمزد",
            "خصوصی/دولتی",
            "مبلغ حق بیمه (با مالیات)",
            "آئین نامه قبل از 1403/03/20",
            "کسر 9% مالیات",
            "کسر 10% مالیات"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border_style
        
        # تنظیم عرض ستون‌ها
        column_widths = [8, 20, 15, 25, 15, 25, 25, 18, 18]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col)].width = width
        
        # اضافه کردن یک ردیف نمونه
        sample_row = [1, "", "", "", "", "", "", "", ""]
        for col, value in enumerate(sample_row, 1):
            cell = ws.cell(row=2, column=col)
            cell.value = value
            cell.border = border_style
            if col == 1:
                cell.alignment = Alignment(horizontal="center")
        
        # راهنما در sheet دوم
        ws2 = wb.create_sheet("راهنما")
        instructions = [
            ["راهنمای استفاده از فایل ورودی"],
            [""],
            ["ستون‌ها:"],
            ["1. کارمزد بازاریابی: درصد کارمزد بازاریابی را وارد کنید (مثال: 5)"],
            ["2. درصد صدور: درصد صدور را وارد کنید (مثال: 3)"],
            ["3. سهم بازاریاب: یکی از گزینه‌های زیر را وارد کنید:"],
            ["   - 20% صدور"],
            ["   - 40% معرف"],
            ["   - 60% صدور و معرف"],
            ["4. خصوصی/دولتی: یکی از گزینه‌های زیر را وارد کنید:"],
            ["   - 100% خصوصی"],
            ["   - 50% دولتی"],
            ["5. مبلغ حق بیمه: مبلغ حق بیمه را وارد کنید (بدون کاما)"],
            ["6. آئین نامه قبل از 1403/03/20: بله یا خیر"],
            ["7. کسر 9% مالیات: بله یا خیر"],
            ["8. کسر 10% مالیات: بله یا خیر"],
            [""],
            ["نکته: بعد از پر کردن اطلاعات، فایل را ذخیره کنید و سپس از منوی 'خروجی اکسل' استفاده کنید."]
        ]
        
        for row, instruction in enumerate(instructions, 1):
            ws2.cell(row=row, column=1, value=instruction[0])
        
        wb.save(file_path)
        global excel_input_path
        excel_input_path = file_path
        
        # باز کردن فایل
        open_excel_file(file_path)
        
        messagebox.showinfo("موفق", f"فایل template ایجاد شد و باز شد!\n\nمسیر: {file_path}\n\nلطفا اطلاعات را وارد کرده و فایل را ذخیره کنید.")
        return file_path
    except Exception as e:
        messagebox.showerror("خطا", f"خطا در ایجاد فایل: {str(e)}")
        return None

def create_sample_excel(): #ایجاد فایل نمونه اکسل با داده‌های نمونه
    """ایجاد فایل نمونه اکسل با داده‌های نمونه برای راهنمایی کاربر"""
    global HAS_OPENPYXL, Workbook, load_workbook, Font, PatternFill, Alignment, Border, Side
    if not HAS_OPENPYXL:
        result = messagebox.askyesno(
            "کتابخانه مورد نیاز", 
            "کتابخانه openpyxl برای کار با فایل‌های اکسل نیاز است.\n\n"
            "آیا می‌خواهید اکنون نصب شود؟\n\n"
            "(نیاز به دسترسی به اینترنت)"
        )
        if result:
            try:
                import subprocess
                import sys
                subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
                messagebox.showinfo("موفق", "کتابخانه openpyxl با موفقیت نصب شد!\nلطفا دوباره تلاش کنید.")
                try:
                    global Workbook, load_workbook, Font, PatternFill, Alignment, Border, Side
                    from openpyxl import Workbook, load_workbook
                    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                    HAS_OPENPYXL = True
                except Exception:
                    pass
            except Exception as e:
                messagebox.showerror("خطا", f"خطا در نصب کتابخانه:\n{str(e)}\n\nلطفا دستی نصب کنید:\npip install openpyxl")
        return None
    
    try:
        # انتخاب مسیر ذخیره
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            title="ذخیره فایل نمونه اکسل",
            initialfile="نمونه_ورودی_اکسل.xlsx"
        )
        
        if not file_path:
            return None
        
        wb = Workbook()
        ws = wb.active
        ws.title = "ورودی"
        
        # استایل برای هدر
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")
        border_style = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # نوشتن هدرها
        headers = [
            "ردیف",
            "کارمزد بازاریابی (%)",
            "درصد صدور (%)",
            "سهم بازاریاب از کارمزد",
            "خصوصی/دولتی",
            "مبلغ حق بیمه (با مالیات)",
            "آئین نامه قبل از 1403/03/20",
            "کسر 9% مالیات",
            "کسر 10% مالیات"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border_style
        
        # داده‌های نمونه
        sample_data = [
            [1, 5, 3, "20% صدور", "100% خصوصی", 50000000, "خیر", "خیر", "خیر"],
            [2, 4.5, 2.5, "40% معرف", "100% خصوصی", 20000000000, "خیر", "بله", "خیر"],
            [3, 6, 3.5, "60% صدور و معرف", "50% دولتی", 35000000000, "خیر", "خیر", "بله"],
            [4, 5.5, 3, "20% صدور", "100% خصوصی", 12000000000, "بله", "خیر", "خیر"],
            [5, 4, 2, "40% معرف", "50% دولتی", 8000000000, "خیر", "خیر", "خیر"],
        ]
        
        # نوشتن داده‌های نمونه
        for row_idx, row_data in enumerate(sample_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.border = border_style
                if col_idx == 1:
                    cell.alignment = Alignment(horizontal="center")
                elif col_idx in [2, 3, 6]:
                    cell.alignment = Alignment(horizontal="center")
                elif col_idx in [7, 8, 9]:
                    cell.alignment = Alignment(horizontal="center")
        
        # تنظیم عرض ستون‌ها
        column_widths = [8, 20, 15, 25, 15, 25, 25, 18, 18]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col)].width = width
        
        # راهنما در sheet دوم
        ws2 = wb.create_sheet("راهنما")
        instructions = [
            ["راهنمای استفاده از فایل ورودی"],
            [""],
            ["این فایل شامل 5 ردیف نمونه است که می‌توانید آن‌ها را تغییر دهید یا ردیف‌های جدید اضافه کنید."],
            [""],
            ["ستون‌ها:"],
            ["1. کارمزد بازاریابی: درصد کارمزد بازاریابی را وارد کنید (مثال: 5)"],
            ["2. درصد صدور: درصد صدور را وارد کنید (مثال: 3)"],
            ["3. سهم بازاریاب: یکی از گزینه‌های زیر را وارد کنید:"],
            ["   - 20% صدور"],
            ["   - 40% معرف"],
            ["   - 60% صدور و معرف"],
            ["4. خصوصی/دولتی: یکی از گزینه‌های زیر را وارد کنید:"],
            ["   - 100% خصوصی"],
            ["   - 50% دولتی"],
            ["5. مبلغ حق بیمه: مبلغ حق بیمه را وارد کنید (بدون کاما)"],
            ["6. آئین نامه قبل از 1403/03/20: بله یا خیر"],
            ["7. کسر 9% مالیات: بله یا خیر"],
            ["8. کسر 10% مالیات: بله یا خیر"],
            [""],
            ["نکات مهم:"],
            ["- بعد از پر کردن اطلاعات، فایل را ذخیره کنید"],
            ["- سپس از منوی 'فایل ها' → 'خروجی اکسل' استفاده کنید"],
            ["- می‌توانید ردیف‌های نمونه را پاک کرده و داده‌های خود را وارد کنید"],
            ["- می‌توانید ردیف‌های جدید اضافه کنید"]
        ]
        
        for row, instruction in enumerate(instructions, 1):
            cell = ws2.cell(row=row, column=1, value=instruction[0])
            if row <= 2:
                cell.font = Font(bold=True, size=12)
        
        # تنظیم عرض ستون راهنما
        ws2.column_dimensions['A'].width = 80
        
        wb.save(file_path)
        
        # باز کردن فایل
        open_excel_file(file_path)
        
        messagebox.showinfo("موفق", 
            f"فایل نمونه با موفقیت ایجاد شد!\n\n"
            f"مسیر: {file_path}\n\n"
            f"این فایل شامل 5 ردیف نمونه است.\n"
            f"می‌توانید داده‌ها را تغییر دهید یا ردیف‌های جدید اضافه کنید.")
        return file_path
    except Exception as e:
        messagebox.showerror("خطا", f"خطا در ایجاد فایل نمونه: {str(e)}")
        return None

def process_excel_output(): #پردازش و ایجاد فایل خروجی اکسل
    """خواندن فایل ورودی و ایجاد فایل خروجی با نتایج"""
    global HAS_OPENPYXL, Workbook, load_workbook, Font, PatternFill, Alignment, Border, Side, excel_input_path
    if not HAS_OPENPYXL:
        result = messagebox.askyesno(
            "کتابخانه مورد نیاز", 
            "کتابخانه openpyxl برای کار با فایل‌های اکسل نیاز است.\n\n"
            "آیا می‌خواهید اکنون نصب شود؟\n\n"
            "(نیاز به دسترسی به اینترنت)"
        )
        if result:
            try:
                import subprocess
                import sys
                subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
                messagebox.showinfo("موفق", "کتابخانه openpyxl با موفقیت نصب شد!\nلطفا دوباره تلاش کنید.")
                # reload imports
                try:
                    from openpyxl import Workbook, load_workbook
                    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                    HAS_OPENPYXL = True
                except:
                    pass
            except Exception as e:
                messagebox.showerror("خطا", f"خطا در نصب کتابخانه:\n{str(e)}\n\nلطفا دستی نصب کنید:\npip install openpyxl")
        return
    
    global excel_input_path
    
    # اگر فایل ورودی انتخاب نشده، از کاربر بپرس
    if not excel_input_path or not os.path.exists(excel_input_path):
        excel_input_path = filedialog.askopenfilename(
            title="انتخاب فایل ورودی اکسل",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
    
    if not excel_input_path:
        messagebox.showwarning("هشدار", "فایل ورودی انتخاب نشد!")
        return
    
    try:
        # خواندن فایل ورودی
        wb = load_workbook(excel_input_path)
        ws = wb.active
        
        # بررسی وجود داده
        if ws.max_row < 2:
            messagebox.showwarning("هشدار", "فایل ورودی خالی است یا فقط هدر دارد!")
            return

        # خواندن هدرها و نگاشت ستون‌ها به نام‌های منطقی.
        # این باعث می‌شود قالب‌های مختلف (مثلاً فایل شما `template.xlsx` با 6 ستون)
        # و قالب قدیمی 9 ستونی پشتیبانی شوند.
        header_map = {}
        headers = []
        for col_idx in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=col_idx).value
            h = str(val).strip() if val is not None else ""
            headers.append(h)
            low = h.lower()
            if 'بازاریابی' in low and 'کارمزد' in low:
                header_map['bazaryabi_prc'] = col_idx
            elif ('صدور' in low and ('کارمزد' in low or '%' in low)) or low == 'کارمزد صدور':
                header_map['expert_prc'] = col_idx
            elif 'سهم' in low and 'بازاریاب' in low:
                header_map['share_option'] = col_idx
            elif 'خصوصی' in low or 'دولتی' in low or 'درصد خصوصی' in low:
                header_map['private_option'] = col_idx
            elif 'حق بیمه' in low:
                header_map['ins_cost'] = col_idx
            elif 'آئین' in low or 'قبل' in low:
                header_map['old_regulation'] = col_idx
            elif '9%' in low or '9' in low and 'مالیات' in low:
                header_map['tax_9'] = col_idx
            elif '10%' in low or '10' in low and 'مالیات' in low:
                header_map['tax_10'] = col_idx
            elif 'مالیات' in low:
                # generic tax percent column (e.g., "درصد مالیات")
                header_map['tax_percent'] = col_idx

        # اگر نگاشت پیدا نشد، از مکان‌های پیش‌فرض قبلی استفاده کن
        def col(name, default):
            return header_map.get(name, default)

        
        # انتخاب مسیر ذخیره فایل خروجی (پیش‌فرض نام: output.xlsx)
        output_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            initialfile="output.xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            title="ذخیره فایل خروجی اکسل"
        )
        
        if not output_path:
            return
        
        # ایجاد workbook جدید برای خروجی
        wb_output = Workbook()
        ws_output = wb_output.active
        ws_output.title = "نتایج"
        
        # استایل‌ها
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        result_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
        result_font = Font(bold=True, color="10B981", size=11)
        border_style = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal="center", vertical="center")
        
        # نوشتن هدرهای خروجی
        # Build list of input columns to copy to output, skipping the
        # old-regulation column (آئین نامه قبل از 1403/03/20) if present.
        input_cols = []
        input_headers = []
        for c in range(1, ws.max_column + 1):
            h = ws.cell(row=1, column=c).value
            h_str = str(h).strip() if h is not None else ""
            # skip the old-regulation column in output
            if 'آئین' in h_str and '1403' in h_str:
                continue
            # skip any existing plain 'کارمزد نهایی' column (keep only 'کارمزد نهایی (ریال)')
            if h_str == 'کارمزد نهایی' or (h_str.startswith('کارمزد نهایی') and '(ریال)' not in h_str):
                continue
            input_cols.append(c)
            input_headers.append(h_str)

        # write input headers to output
        for idx, h in enumerate(input_headers, 1):
            cell = ws_output.cell(row=1, column=idx)
            cell.value = h
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = border_style

        # result columns come immediately after copied input columns
        result_col_1 = len(input_headers) + 1
        result_col_2 = result_col_1 + 1

        ws_output.cell(row=1, column=result_col_1, value="کارمزد نهایی (ریال)").fill = header_fill
        ws_output.cell(row=1, column=result_col_1).font = header_font
        ws_output.cell(row=1, column=result_col_1).alignment = center_alignment
        ws_output.cell(row=1, column=result_col_1).border = border_style

        ws_output.cell(row=1, column=result_col_2, value="درصد کارمزد نهایی").fill = header_fill
        ws_output.cell(row=1, column=result_col_2).font = header_font
        ws_output.cell(row=1, column=result_col_2).alignment = center_alignment
        ws_output.cell(row=1, column=result_col_2).border = border_style
        
        # پردازش هر ردیف
        success_count = 0
        error_count = 0
        
        for row_idx in range(2, ws.max_row + 1):
            try:
                # خواندن داده‌ها
                # خواندن با استفاده از نگاشت هدرها (سازگار با چند قالب)
                row_num = ws.cell(row=row_idx, column=col('row_num', 1)).value
                bazaryabi_prc = ws.cell(row=row_idx, column=col('bazaryabi_prc', 2)).value
                expert_prc = ws.cell(row=row_idx, column=col('expert_prc', 3)).value
                share_option = ws.cell(row=row_idx, column=col('share_option', 4)).value or ""
                private_option = ws.cell(row=row_idx, column=col('private_option', 5)).value or ""
                ins_cost = ws.cell(row=row_idx, column=col('ins_cost', 6)).value
                old_regulation = ws.cell(row=row_idx, column=col('old_regulation', 7)).value
                tax_9 = ws.cell(row=row_idx, column=col('tax_9', 8)).value
                tax_10 = ws.cell(row=row_idx, column=col('tax_10', 9)).value
                tax_percent = ws.cell(row=row_idx, column=header_map.get('tax_percent', 0)).value if header_map.get('tax_percent') else None

                # کپی داده‌های ورودی (فقط ستون‌های انتخاب شده)
                for out_i, c in enumerate(input_cols, 1):
                    cell_value = ws.cell(row=row_idx, column=c).value
                    ws_output.cell(row=row_idx, column=out_i, value=cell_value)
                    ws_output.cell(row=row_idx, column=out_i).border = border_style
                
                # تبدیل مقادیر
                if bazaryabi_prc is None or expert_prc is None or ins_cost is None:
                    ws_output.cell(row=row_idx, column=10, value="داده ناقص")
                    ws_output.cell(row=row_idx, column=11, value="داده ناقص")
                    error_count += 1
                    continue
                
                # تبدیل مبلغ حق بیمه
                if isinstance(ins_cost, str):
                    ins_cost = ins_cost.replace(',', '').replace(' ', '')
                try:
                    cost = float(ins_cost)
                except:
                    ws_output.cell(row=row_idx, column=10, value="خطا در مبلغ")
                    ws_output.cell(row=row_idx, column=11, value="خطا در مبلغ")
                    error_count += 1
                    continue
                
                # تنظیم متغیرهای پلکان بر اساس آئین نامه (استفاده از متغیرهای محلی)
                old_reg = str(old_regulation).lower() in ['بله', 'yes', 'y', '1', 'true']
                if old_reg:
                    local_bzy1, local_bzy2, local_bzy3, local_bzy4 = 0, 12500000000, 25000000000, 50000000000
                    local_expt1, local_expt2, local_expt3, local_expt4 = 0, 2500000000, 12500000000, 25000000000
                else:
                    # استفاده از مقادیر global برای آئین نامه جدید
                    local_bzy1, local_bzy2, local_bzy3, local_bzy4 = bzyLevel1, bzyLevel2, bzyLevel3, bzyLevel4
                    local_expt1, local_expt2, local_expt3, local_expt4 = exptlevel1, exptlevel2, exptlevel3, exptlevel4
                
                # اعمال کسر مالیات
                tax_9_val = False
                tax_10_val = False
                # پشتیبانی از دو حالت: ستون‌های جداگانه 9%/10% یا یک ستون درصد مالیات
                if tax_percent is not None:
                    try:
                        tp = float(str(tax_percent).replace('%', '').replace(',', '').strip())
                        if abs(tp - 9) < 0.001:
                            tax_9_val = True
                        elif abs(tp - 10) < 0.001:
                            tax_10_val = True
                        else:
                            # هر درصد دیگری -> اعمال نسبی
                            cost *= (1 - tp/100)
                    except:
                        # fallback: اگر مقدار غیرقابل‌تبدیلی باشد، چک قدیمی را هم امتحان کن
                        tax_9_val = str(tax_9).lower() in ['بله', 'yes', 'y', '1', 'true']
                        tax_10_val = str(tax_10).lower() in ['بله', 'yes', 'y', '1', 'true']
                else:
                    tax_9_val = str(tax_9).lower() in ['بله', 'yes', 'y', '1', 'true']
                    tax_10_val = str(tax_10).lower() in ['بله', 'yes', 'y', '1', 'true']

                if tax_9_val:
                    cost *= 0.91
                elif tax_10_val:
                    cost *= 0.90
                
                # محاسبه کارمزد بازاریابی
                percent1 = float(bazaryabi_prc)
                result1 = 0.0
                
                if cost > local_bzy1 and cost <= local_bzy2:
                    result1 += (cost * (percent1 * 1))/100
                elif cost > local_bzy2 and cost <= local_bzy3:
                    result1 += (local_bzy2 * (percent1 * 1))/100
                    result1 += ((cost - local_bzy2) * (percent1 * 0.5))/100
                elif cost > local_bzy3 and cost <= local_bzy4:
                    result1 += (local_bzy2 * (percent1 * 1))/100
                    result1 += ((local_bzy3 - local_bzy2) * (percent1 * 0.5))/100
                    result1 += ((cost - local_bzy3) * (percent1 * 0.25))/100
                elif cost > local_bzy4:
                    result1 += (local_bzy2 * (percent1 * 1))/100
                    result1 += ((local_bzy3 - local_bzy2) * (percent1 * 0.5))/100
                    result1 += ((local_bzy4 - local_bzy3) * (percent1 * 0.25))/100
                    result1 += ((cost - local_bzy4) * (percent1 * 0.10))/100
                
                # محاسبه کارمزد صدور
                percent2 = float(expert_prc)
                result2 = 0.0
                
                if cost > local_expt1 and cost <= local_expt2:
                    result2 += (cost * (percent2 * 1))/100
                elif cost > local_expt2 and cost <= local_expt3:
                    result2 += (local_expt2 * (percent2 * 1))/100
                    result2 += ((cost - local_expt2) * (percent2 * 0.25))/100
                elif cost > local_expt3 and cost <= local_expt4:
                    result2 += (local_expt2 * (percent2 * 1))/100
                    result2 += ((local_expt3 - local_expt2) * (percent2 * 0.25))/100
                    result2 += ((cost - local_expt3) * (percent2 * 0.10))/100
                elif cost > local_expt4:
                    result2 += (local_expt2 * (percent2 * 1))/100
                    result2 += ((local_expt3 - local_expt2) * (percent2 * 0.25))/100
                    result2 += ((local_expt4 - local_expt3) * (percent2 * 0.10))/100
                    result2 += ((cost - local_expt4) * (percent2 * 0.5))/100
                
                # محاسبه نهایی
                result = float(result1 + result2)
                
                # اعمال سهم بازاریاب
                share_str = str(share_option).strip()
                if "20" in share_str or "صدور" in share_str:
                    result *= 0.2
                elif "40" in share_str or "معرف" in share_str:
                    result *= 0.4
                elif "60" in share_str:
                    result *= 0.6
                
                # اعمال خصوصی/دولتی
                private_str = str(private_option).strip()
                if "50" in private_str or "دولتی" in private_str:
                    result *= 0.5
                elif "100" in private_str or "خصوصی" in private_str:
                    result *= 1
                
                # محاسبه درصد
                original_cost = float(ins_cost) if isinstance(ins_cost, (int, float)) else float(str(ins_cost).replace(',', ''))
                percentage = (result * 100) / original_cost if original_cost > 0 else 0
                percentage = round(percentage, 2)
                
                # نوشتن نتایج در ستون‌های بعد از ورودی
                ws_output.cell(row=row_idx, column=result_col_1, value=result).number_format = '#,##0'
                ws_output.cell(row=row_idx, column=result_col_1).fill = result_fill
                ws_output.cell(row=row_idx, column=result_col_1).font = result_font
                ws_output.cell(row=row_idx, column=result_col_1).border = border_style

                ws_output.cell(row=row_idx, column=result_col_2, value=f"{percentage}%")
                ws_output.cell(row=row_idx, column=result_col_2).fill = result_fill
                ws_output.cell(row=row_idx, column=result_col_2).font = result_font
                ws_output.cell(row=row_idx, column=result_col_2).border = border_style
                
                success_count += 1
                
            except Exception as e:
                ws_output.cell(row=row_idx, column=result_col_1, value=f"خطا: {str(e)}")
                ws_output.cell(row=row_idx, column=result_col_2, value="خطا")
                error_count += 1
        
        # تنظیم عرض ستون‌ها
        column_widths = [8, 20, 15, 25, 15, 25, 25, 18, 18, 25, 20]
        for col, width in enumerate(column_widths, 1):
            ws_output.column_dimensions[chr(64 + col)].width = width
        
        # ذخیره فایل
        wb_output.save(output_path)
        
        messagebox.showinfo("موفق", 
            f"فایل خروجی با موفقیت ایجاد شد!\n\n"
            f"مسیر: {output_path}\n\n"
            f"تعداد ردیف‌های پردازش شده: {success_count + error_count}\n"
            f"موفق: {success_count}\n"
            f"خطا: {error_count}")
        
        # باز کردن فایل خروجی
        open_excel_file(output_path)
        
    except Exception as e:
        messagebox.showerror("خطا", f"خطا در پردازش فایل: {str(e)}")

def format_number(value): #فرمت کردن عدد با جداکننده هزارگان
    """تبدیل عدد به رشته با جداکننده کاما"""
    try:
        # حذف کاماها و فاصله‌ها
        num_str = str(value).replace(',', '').replace(' ', '')
        if num_str == '':
            return ''
        num = int(num_str)
        return f"{num:,}"
    except:
        return str(value)

def format_entry_number(entry_widget): #فرمت کردن Entry بعد از تایپ
    """فرمت کردن عدد در Entry با حفظ موقعیت مکان‌نما"""
    current = entry_widget.get()
    cursor_pos = entry_widget.index(INSERT)
    
    # حذف کاماها و فاصله‌ها
    text_without_commas = current.replace(',', '').replace(' ', '')
    
    # اگر خالی است، برگردان
    if not text_without_commas:
        return
    
    # شمارش ارقام قبل از مکان‌نما
    digits_before_cursor = 0
    for i in range(min(cursor_pos, len(current))):
        if current[i].isdigit():
            digits_before_cursor += 1
    
    # فرمت کردن
    try:
        num = int(text_without_commas)
        formatted = f"{num:,}"
    except:
        formatted = text_without_commas
    
    # تنظیم مجدد متن
    entry_widget.delete(0, END)
    entry_widget.insert(0, formatted)
    
    # تنظیم موقعیت مکان‌نما
    if formatted and digits_before_cursor > 0:
        new_cursor = 0
        digit_count = 0
        for i, char in enumerate(formatted):
            if char.isdigit():
                digit_count += 1
                if digit_count >= digits_before_cursor:
                    new_cursor = i + 1
                    break
        entry_widget.icursor(new_cursor)
    else:
        entry_widget.icursor(END)

def get_number_value(entry_widget): #گرفتن مقدار عددی از Entry (بدون کاما)
    """گرفتن مقدار عددی از Entry با حذف کاماها"""
    try:
        value = entry_widget.get().replace(',', '').replace(' ', '')
        return int(value) if value else 0
    except:
        return 0
  
def update_values(): #بازنویسی اطلاعات پلکان ها
    global bzyLevel1, bzyLevel2, bzyLevel3, bzyLevel4, exptlevel1, exptlevel2, exptlevel3, exptlevel4
    if checkvar.get():
        bzyLevel1, bzyLevel2, bzyLevel3, bzyLevel4, exptlevel1, exptlevel2, exptlevel3, exptlevel4 = 0, 12500000000, 25000000000, 50000000000, 0, 2500000000, 12500000000 ,25000000000
    else:
        bzyLevel1, bzyLevel2, bzyLevel3, bzyLevel4, exptlevel1, exptlevel2, exptlevel3, exptlevel4 = 0, 15000000000, 30000000000, 60000000000, 0, 3000000000, 15000000000 ,30000000000

class calculation1: #فرمول محسابه درصد بازاریابی
    def __init__(self):
        self.result = 0.0

    def calculate(self):
        try:
            percent1 = float(entryBazaryabiPrc.get().replace(',', '').replace(' ', '') or 0)
            cost = get_number_value(entryInsCost)
            result = 0.0
            
            if checkvar1.get():
                cost *= 0.91  # کاهش 9 درصدی
            elif checkvar2.get():
                cost *= 0.90  # کاهش 10 درصدی
        
            if cost > bzyLevel1 and cost <= bzyLevel2:
                result += (cost * (percent1 * 1))/100
                
            elif cost > bzyLevel2 and cost <= bzyLevel3:
                result += (bzyLevel2 * (percent1 * 1))/100
                result += ((cost - bzyLevel2) * (percent1 * 0.5))/100
                
            elif cost > bzyLevel3 and cost <= bzyLevel4:
                result += (bzyLevel2 * (percent1 * 1))/100
                result += ((bzyLevel3 - bzyLevel2) * (percent1 * 0.5))/100
                result += ((cost - bzyLevel3) * (percent1 * 0.25))/100
                
            elif cost > bzyLevel4:
                result += (bzyLevel2 * (percent1 * 1))/100
                result += ((bzyLevel3 - bzyLevel2) * (percent1 * 0.5))/100
                result += ((bzyLevel4 - bzyLevel3) * (percent1 * 0.25))/100
                result += ((cost - bzyLevel4) * (percent1 * 0.10))/100
                
            return result
        except Exception as e:
            print(f"Error in calculation1: {e}")
            return 0.0
            
class calculation2: #فرمول محاسبه درصد کارمزد
    def __init__(self):
        self.result = 0.0

    def calculate(self):
        try:
            percent2 = float(entryExpertPrc.get().replace(',', '').replace(' ', '') or 0)
            cost = get_number_value(entryInsCost)
            result = 0.0

            if checkvar1.get():
                cost *= 0.91  # کاهش 9 درصدی
            elif checkvar2.get():
                cost *= 0.90  # کاهش 10 درصدی
        
            if cost > exptlevel1 and cost <= exptlevel2:
                result += (cost * (percent2 * 1))/100
                    
            elif cost > exptlevel2 and cost <= exptlevel3:
                result += (exptlevel2 * (percent2 * 1))/100
                result += ((cost - exptlevel2) * (percent2 * 0.25))/100
                        
            elif cost > exptlevel3 and cost <= exptlevel4:
                result += (exptlevel2 * (percent2 * 1))/100
                result += ((exptlevel3 - exptlevel2) * (percent2 * 0.25))/100
                result += ((cost - exptlevel3) * (percent2 * 0.10))/100
                    
            elif cost > exptlevel4:
                result += (exptlevel2 * (percent2 * 1))/100
                result += ((exptlevel3 - exptlevel2) * (percent2 * 0.25))/100
                result += ((exptlevel4 - exptlevel3) * (percent2 * 0.10))/100
                result += ((cost - exptlevel4) * (percent2 * 0.5))/100
                
            return result
        except Exception as e:
            print(f"Error in calculation2: {e}")
            return 0.0
                             
def calculate(): #محاسبه نهایی کارمزد
    try:
        
        clear_error_message()

        calc1 = calculation1().calculate()
        calc2 = calculation2().calculate()
        cost = get_number_value(entryInsCost)
        
        if checkvar1.get():
            cost *= 0.91  # کاهش 9 درصدی
        elif checkvar2.get():
            cost *= 0.90  # کاهش 10 درصدی

        result = float(calc1 + calc2)

        selected_option1 = clicked1.get()
        if selected_option1 == "%20 صدور":
            result *= 0.2
        elif selected_option1 == "%40 معرف":
            result *= 0.4
        elif selected_option1 == "%60 صدور و معرف":
            result *= 0.6

        selected_option2 = clicked2.get()
        if selected_option2 == "%100 خصوصی":
            result *= 1
        elif selected_option2 == "%50 دولتی":
            result *= 0.5

        krbldFinalPercentage = float((result * 100)/cost)
        x = round(krbldFinalPercentage , 2) # باعث نمایش مقدار رند شده با دو رقم اعشار می شود.
        
        labelErrors.config(text=f"درصد کارمزد نهایی به کل حق بیمه: {x}%", fg=COLORS['success'])
        labelFinalKarmozd.config(text=f"{result:,.0f} ریال", fg=COLORS['success'])
        return result
    except Exception as e:
        labelErrors.config(text=f"خطا: {str(e)}", fg=COLORS['error'])
#endregion
#region ///////////////////////////////////////////////////////////// menu bar
menubar = Menu(root, bg=COLORS['card'], fg=COLORS['text'], 
               activebackground=COLORS['accent'], activeforeground='white',
               font=font_small, tearoff=0)

file = Menu(menubar, tearoff=0, bg=COLORS['card'], fg=COLORS['text'],
            activebackground=COLORS['accent'], activeforeground='white') 
menubar.add_cascade(label='فایل ها', menu=file) 
file.add_command(label='نمونه فایل', command=download_template_file)
file.add_separator()
file.add_command(label='خروج', command=root.destroy) 

download = Menu(menubar, tearoff=0, bg=COLORS['card'], fg=COLORS['text'],
               activebackground=COLORS['accent'], activeforeground='white') 
menubar.add_cascade(label='دانلودها', menu=download) 
download.add_command(label='آئین نامه 102', command=openWeb1) 
download.add_command(label='آئین نامه 102.1', command=openWeb2) 

help_ = Menu(menubar, tearoff=0, bg=COLORS['card'], fg=COLORS['text'],
            activebackground=COLORS['accent'], activeforeground='white')
menubar.add_cascade(label='راهنما', menu=help_) 
help_.add_command(label='راهنمایی', command=showInfo)
help_.add_separator() 
help_.add_command(label='درباره ما', command=aboutUs) 
#endregion

#region ///////////////////////////////////////////////////////////// Header Section
header_frame = Frame(root, bg=COLORS['header'], height=80)
header_frame.grid(row=0, column=0, sticky=EW, padx=0, pady=0)
header_frame.grid_propagate(False)
header_frame.grid_columnconfigure(0, weight=1)

title_label = Label(header_frame, text="محاسبه کارمزد پلکانی", 
                   font=font_title, bg=COLORS['header'], fg='white')
title_label.grid(row=0, column=0, pady=20)

subtitle_label = Label(header_frame, text="طبق آئین نامه بیمه مرکزی", 
                      font=font_small, bg=COLORS['header'], fg=COLORS['text_light'])
subtitle_label.grid(row=1, column=0, pady=(0, 15))
#endregion

#region ///////////////////////////////////////////////////////////// Scrollable Container
#region ///////////////////////////////////////////////////////////// Main Container
main_container = Frame(root, bg=COLORS['bg'], padx=padding*2, pady=padding*2)
main_container.grid(row=1, column=0, sticky=NSEW, padx=0, pady=0)
main_container.grid_columnconfigure(0, weight=1)
main_container.grid_rowconfigure(0, weight=0)  # Input card
main_container.grid_rowconfigure(1, weight=0)  # Options card
main_container.grid_rowconfigure(2, weight=0)  # Button
main_container.grid_rowconfigure(3, weight=1)  # Result card - expandable

#region ////////////////////////////// Input Card
input_card = Frame(main_container, bg=COLORS['card'], relief=FLAT,
                  highlightbackground=COLORS['border'], highlightthickness=1)
input_card.grid(row=0, column=0, sticky=EW, pady=(0, padding))
input_card.grid_columnconfigure(1, weight=1, minsize=200)

card_inner = Frame(input_card, bg=COLORS['card'], padx=inner_padding*2, pady=inner_padding*2)
card_inner.grid(row=0, column=0, columnspan=2, sticky=EW)
card_inner.grid_columnconfigure(1, weight=1)

# Title for input card
card_title = Label(card_inner, text="اطلاعات ورودی", font=font_heading,
                  bg=COLORS['card'], fg=COLORS['text'])
card_title.grid(row=0, column=0, columnspan=2, sticky=W, pady=(0, inner_padding*2))

# Separator
separator1 = Frame(card_inner, bg=COLORS['border_light'], height=1)
separator1.grid(row=1, column=0, columnspan=2, sticky=EW, pady=(0, inner_padding*2))

# Input Fields
labelBazaryabiPrc = Label(card_inner, text="کارمزد بازاریابی (%)", font=font_heading, 
                         bg=COLORS['card'], fg=COLORS['text_secondary'])
labelBazaryabiPrc.grid(row=2, column=0, sticky=W, padx=(0, inner_padding), pady=inner_padding)

entryBazaryabiPrc = Entry(card_inner, justify=CENTER, font=font_body,
                          relief=SOLID, borderwidth=1, highlightthickness=2,
                          highlightbackground=COLORS['border'], highlightcolor=COLORS['accent'],
                          bg=COLORS['card'], fg=COLORS['text'], insertbackground=COLORS['accent'])
entryBazaryabiPrc.grid(row=2, column=1, sticky=EW, padx=(0, 0), pady=inner_padding)

labelExpertPrc = Label(card_inner, text="درصد صدور (%)", font=font_heading,
                       bg=COLORS['card'], fg=COLORS['text_secondary'])
labelExpertPrc.grid(row=3, column=0, sticky=W, padx=(0, inner_padding), pady=inner_padding)

entryExpertPrc = Entry(card_inner, justify=CENTER, font=font_body,
                       relief=SOLID, borderwidth=1, highlightthickness=2,
                       highlightbackground=COLORS['border'], highlightcolor=COLORS['accent'],
                       bg=COLORS['card'], fg=COLORS['text'], insertbackground=COLORS['accent'])
entryExpertPrc.grid(row=3, column=1, sticky=EW, padx=(0, 0), pady=inner_padding)

options1 = ["انتخاب کنید", "%20 صدور", "%40 معرف", "%60 صدور و معرف"]
clicked1 = StringVar() 
clicked1.set("انتخاب کنید")

labelExpertMood1 = Label(card_inner, text="سهم بازاریاب از کارمزد", font=font_heading,
                        bg=COLORS['card'], fg=COLORS['text_secondary'])
labelExpertMood1.grid(row=4, column=0, sticky=W, padx=(0, inner_padding), pady=inner_padding)

drop1 = ttk.Combobox(card_inner, textvariable=clicked1, values=options1, 
                    state="readonly", font=font_body)
drop1.grid(row=4, column=1, sticky=EW, padx=(0, 0), pady=inner_padding)

# Style ttk combobox
style = ttk.Style()
style.theme_use('clam')
style.configure('TCombobox', fieldbackground=COLORS['card'], borderwidth=1,
                relief=SOLID, padding=5)
style.map('TCombobox', fieldbackground=[('readonly', COLORS['card'])],
         selectbackground=[('readonly', COLORS['accent_light'])],
         selectforeground=[('readonly', COLORS['text'])])

options2 = ["انتخاب کنید", "%100 خصوصی", "%50 دولتی"]
clicked2 = StringVar() 
clicked2.set("انتخاب کنید") 

labelExpertMood = Label(card_inner, text="خصوصی/دولتی", font=font_heading,
                       bg=COLORS['card'], fg=COLORS['text_secondary'])
labelExpertMood.grid(row=5, column=0, sticky=W, padx=(0, inner_padding), pady=inner_padding)

drop2 = ttk.Combobox(card_inner, textvariable=clicked2, values=options2, 
                    state="readonly", font=font_body)
drop2.grid(row=5, column=1, sticky=EW, padx=(0, 0), pady=inner_padding)

labelInsCost = Label(card_inner, text="مبلغ حق بیمه (با مالیات)", font=font_heading,
                    bg=COLORS['card'], fg=COLORS['text_secondary'])
labelInsCost.grid(row=6, column=0, sticky=W, padx=(0, inner_padding), pady=inner_padding)

entryInsCost = Entry(card_inner, justify=CENTER, font=font_body,
                    relief=SOLID, borderwidth=1, highlightthickness=2,
                    highlightbackground=COLORS['border'], highlightcolor=COLORS['accent'],
                    bg=COLORS['card'], fg=COLORS['text'], insertbackground=COLORS['accent'])
entryInsCost.grid(row=6, column=1, sticky=EW, padx=(0, 0), pady=inner_padding)
# فرمت کردن عدد هنگام تایپ
def on_ins_cost_key(event):
    """مدیریت ورود عدد در فیلد مبلغ حق بیمه"""
    if event.keysym in ['BackSpace', 'Delete', 'Left', 'Right', 'Up', 'Down', 'Home', 'End', 'Tab']:
        return
    if event.state & 0x4 or event.state & 0x1:
        return
    if event.char and not event.char.isdigit():
        return 'break'
    root.after_idle(lambda: format_entry_number(entryInsCost))

entryInsCost.bind('<KeyPress>', on_ins_cost_key)
#endregion

#region ////////////////////////////// Options Card
options_card = Frame(main_container, bg=COLORS['card'], relief=FLAT,
                    highlightbackground=COLORS['border'], highlightthickness=1)
options_card.grid(row=1, column=0, sticky=EW, pady=(0, padding))
options_card.grid_columnconfigure(0, weight=1)
options_card.grid_columnconfigure(1, weight=1)

options_inner = Frame(options_card, bg=COLORS['card'], padx=inner_padding*2, pady=inner_padding*2)
options_inner.grid(row=0, column=0, columnspan=2, sticky=EW)
options_inner.grid_columnconfigure(0, weight=1)
options_inner.grid_columnconfigure(1, weight=1)

LabelFrame1 = Frame(options_inner, bg=COLORS['card'])
LabelFrame1.grid(row=0, column=0, sticky=EW, padx=(0, padding//2))
LabelFrame1.grid_columnconfigure(0, weight=1)

checkvar = BooleanVar()

# تابع آپلود فایل اکسل
def upload_excel_file(): #آپلود فایل اکسل
    """انتخاب و آپلود فایل اکسل ورودی"""
    global excel_input_path, HAS_OPENPYXL, Workbook, load_workbook, Font, PatternFill, Alignment, Border, Side
    
    if not HAS_OPENPYXL:
        result = messagebox.askyesno(
            "کتابخانه مورد نیاز", 
            "کتابخانه openpyxl برای کار با فایل‌های اکسل نیاز است.\n\n"
            "آیا می‌خواهید اکنون نصب شود؟\n\n"
            "(نیاز به دسترسی به اینترنت)"
        )
        if result:
            try:
                import subprocess
                import sys
                subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
                messagebox.showinfo("موفق", "کتابخانه openpyxl با موفقیت نصب شد!\nلطفا دوباره تلاش کنید.")
                try:
                    global Workbook, load_workbook, Font, PatternFill, Alignment, Border, Side
                    from openpyxl import Workbook, load_workbook
                    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                    HAS_OPENPYXL = True
                except Exception:
                    pass
            except Exception as e:
                messagebox.showerror("خطا", f"خطا در نصب کتابخانه:\n{str(e)}\n\nلطفا دستی نصب کنید:\npip install openpyxl")
        return
    
    try:
        # انتخاب فایل اکسل
        file_path = filedialog.askopenfilename(
            title="انتخاب فایل اکسل ورودی",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        
        if not file_path:
            return
        
        # بررسی وجود فایل
        if not os.path.exists(file_path):
            messagebox.showerror("خطا", "فایل انتخاب شده وجود ندارد!")
            return
        
        # بررسی فرمت فایل
        if not file_path.lower().endswith(('.xlsx', '.xls')):
            messagebox.showerror("خطا", "فایل انتخاب شده یک فایل اکسل معتبر نیست!")
            return
        
        # بررسی محتوای فایل
        try:
            wb = load_workbook(file_path)
            ws = wb.active
            if ws.max_row < 2:
                messagebox.showwarning("هشدار", "فایل انتخاب شده خالی است یا فقط هدر دارد!")
                return
        except Exception as e:
            messagebox.showerror("خطا", f"خطا در خواندن فایل:\n{str(e)}")
            return
        
        # ذخیره مسیر فایل
        excel_input_path = file_path
        
        # نمایش پیام موفقیت
        file_name = os.path.basename(file_path)
        messagebox.showinfo("موفق", 
            f"فایل با موفقیت آپلود شد!\n\n"
            f"نام فایل: {file_name}\n"
            f"تعداد ردیف‌ها: {ws.max_row - 1}\n\n"
            f"اکنون می‌توانید از دکمه 'محاسبه کارمزد فایل اکسل' استفاده کنید.")
        
    except Exception as e:
        messagebox.showerror("خطا", f"خطا در آپلود فایل:\n{str(e)}")

# تابع محاسبه گروهی کارمزد
def calculate_batch_commission(): #محاسبه کارمزد گروهی از فایل اکسل
    """محاسبه کارمزد برای تمام ردیف‌های فایل اکسل"""
    global excel_input_path, HAS_OPENPYXL, Workbook, load_workbook, Font, PatternFill, Alignment, Border, Side
    
    if not HAS_OPENPYXL:
        result = messagebox.askyesno(
            "کتابخانه مورد نیاز", 
            "کتابخانه openpyxl برای کار با فایل‌های اکسل نیاز است.\n\n"
            "آیا می‌خواهید اکنون نصب شود؟\n\n"
            "(نیاز به دسترسی به اینترنت)"
        )
        if result:
            try:
                import subprocess
                import sys
                subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
                messagebox.showinfo("موفق", "کتابخانه openpyxl با موفقیت نصب شد!\nلطفا دوباره تلاش کنید.")
                try:
                    from openpyxl import Workbook, load_workbook
                    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                    HAS_OPENPYXL = True
                except:
                    pass
            except Exception as e:
                messagebox.showerror("خطا", f"خطا در نصب کتابخانه:\n{str(e)}\n\nلطفا دستی نصب کنید:\npip install openpyxl")
        return
    
    # بررسی وجود فایل
    if not excel_input_path or not os.path.exists(excel_input_path):
        messagebox.showwarning("هشدار", "ابتدا فایل اکسل را آپلود کنید!")
        return
    
    # استفاده از تابع موجود process_excel_output
    process_excel_output()

# دکمه آپلود فایل اکسل
upload_button = Button(LabelFrame1, text="📁 آپلود فایل اکسل", font=font_small,
                      bg=COLORS['accent'], fg="white", relief=FLAT,
                      padx=15, pady=8, cursor="hand2", command=upload_excel_file,
                      activebackground=COLORS['accent_hover'], activeforeground='white')
upload_button.grid(row=0, column=0, sticky=EW, padx=inner_padding, pady=inner_padding//2)

# دکمه محاسبه کارمزد فایل اکسل
calculate_excel_button = Button(LabelFrame1, text="⚙️ محاسبه کارمزد فایل اکسل", font=font_small,
                               bg=COLORS['success'], fg="white", relief=FLAT,
                               padx=15, pady=8, cursor="hand2", command=calculate_batch_commission,
                               activebackground='#059669', activeforeground='white')
calculate_excel_button.grid(row=1, column=0, sticky=EW, padx=inner_padding, pady=inner_padding//2)


# Tax checkboxes: place under the insurance amount input inside the input card
checkvar1 = BooleanVar()
checkbox1 = Checkbutton(card_inner, text="کسر 9 درصد مالیات", variable=checkvar1, command=None,
                       bg=COLORS['card'], fg=COLORS['text'], font=font_small,
                       activebackground=COLORS['card'], activeforeground=COLORS['text'],
                       selectcolor=COLORS['accent'])
checkbox1.grid(row=7, column=0, sticky=W, padx=inner_padding, pady=inner_padding//2)

checkvar2 = BooleanVar()
checkbox2 = Checkbutton(card_inner, text="کسر 10 درصد مالیات", variable=checkvar2, command=None,
                       bg=COLORS['card'], fg=COLORS['text'], font=font_small,
                       activebackground=COLORS['card'], activeforeground=COLORS['text'],
                       selectcolor=COLORS['accent'])
checkbox2.grid(row=7, column=1, sticky=W, padx=inner_padding, pady=inner_padding//2)
#endregion

#region ////////////////////////////// Button Section
def on_button_enter(e):
    submitButton.config(bg=COLORS['accent_hover'], relief=FLAT)

def on_button_leave(e):
    submitButton.config(bg=COLORS['accent'], relief=FLAT)

# Place the single-row calculate button under the insurance input inside the input card
submitButton = Button(card_inner, text="محاسبه کارمزد", font=font_button,
                     bg=COLORS['accent'], fg="white", relief=FLAT,
                     padx=40, pady=10, cursor="hand2", command=calculate,
                     activebackground=COLORS['accent_hover'], activeforeground='white')
submitButton.grid(row=8, column=0, columnspan=2, pady=(inner_padding*1, inner_padding*2))
submitButton.bind("<Enter>", on_button_enter)
submitButton.bind("<Leave>", on_button_leave)
#endregion

#region ////////////////////////////// Result Card
result_card = Frame(main_container, bg=COLORS['card'], relief=FLAT,
                   highlightbackground=COLORS['border'], highlightthickness=1)
result_card.grid(row=3, column=0, sticky=NSEW, pady=(0, padding))
result_card.grid_columnconfigure(0, weight=1)

result_inner = Frame(result_card, bg=COLORS['card'], padx=inner_padding*2, pady=inner_padding*2)
result_inner.grid(row=0, column=0, sticky=EW)
result_inner.grid_columnconfigure(1, weight=1)

# Header برای کارت نتیجه
result_header = Frame(result_inner, bg=COLORS['card'])
result_header.grid(row=0, column=0, columnspan=3, sticky=EW, pady=(0, inner_padding))
result_header.grid_columnconfigure(1, weight=1)

labelKarmozd = Label(result_header, text="کارمزد نهایی:", font=font_heading,
                    bg=COLORS['card'], fg=COLORS['text_secondary'])
labelKarmozd.grid(row=0, column=0, sticky=W, padx=(0, inner_padding))

labelFinalKarmozd = Label(result_header, text="0 ریال", font=("Segoe UI", 16, "bold"),
                         bg=COLORS['card'], fg=COLORS['success'])
labelFinalKarmozd.grid(row=0, column=1, sticky=W)

# دکمه کپی
def on_copy_enter(e):
    copyButton.config(bg=COLORS['accent_light'])

def on_copy_leave(e):
    copyButton.config(bg=COLORS['card'])

copyButton = Button(result_header, text="📋 کپی", font=font_small,
                   bg=COLORS['card'], fg=COLORS['accent'], relief=FLAT,
                   padx=12, pady=6, cursor="hand2", command=copy_result,
                   activebackground=COLORS['accent_light'], activeforeground=COLORS['accent'],
                   borderwidth=1, highlightthickness=1,
                   highlightbackground=COLORS['border'], highlightcolor=COLORS['accent'])
copyButton.grid(row=0, column=2, sticky=E, padx=(inner_padding, 0))
copyButton.bind("<Enter>", on_copy_enter)
copyButton.bind("<Leave>", on_copy_leave)

labelErrors = Label(result_inner, text="", font=font_small, bg=COLORS['card'], fg=COLORS['error'])
labelErrors.grid(row=1, column=0, columnspan=3, sticky=W, pady=(0, inner_padding))
#endregion
#endregion
#region ///////////////////////////////////////////////////////////// mainloop
centerWindow()
root.config(menu = menubar) 
root.mainloop()
#endregion
